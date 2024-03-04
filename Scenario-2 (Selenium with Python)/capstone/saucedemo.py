import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

@pytest.fixture(scope="module")
def setup():
    #Setup code to initialize the browser
    driver =webdriver.Chrome()
    driver.maximize_window()
    yield driver
    #Teardown code to close the browser
    driver.quit()
def test_verify_page_title(setup):
    driver = setup
    driver.get("https://www.saucedemo.com/")
    assert "Swag Labs" in driver.title
def test_login_successful(setup):
    driver = setup
    driver.get("https://www.saucedemo.com/")
    #Read data from Excel sheet
    workbook = load_workbook("C:\\Users\\Srivani\\PycharmProjects\\capstone\\test_data_positive.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        username = row[0].value
        password = row[1].value
        #Enter username and password
        driver.find_element(By.ID, "user-name").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        # Click on Submit button
        login_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "login-button")))
        login_button.click()
        # Verify successful login
        assert "Swag Labs" in driver.title
        # Logout
        burger_menu = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "react-burger-menu-btn")))
        burger_menu.click()
        logout_link = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "logout_sidebar_link")))
        logout_link.click()
def test_login_unsuccessful(setup):
    driver = setup
    driver.get("https://www.saucedemo.com/")
    #Read data from Excel sheet
    workbook = load_workbook("C:\\Users\\Srivani\\PycharmProjects\\capstone\\test_data_negative.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        username = row[0].value
        password = row[1].value
        #Enter username and password
        driver.find_element(By.ID, "user-name").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        #Click on Submit button
        driver.find_element(By.ID, "login-button").click()
        #Verify error message
        error_message = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "//h3[@data-test='error']")))
        assert error_message.text == "Epic sadface: Username and password do not match any user in this service"
        # Clear input fields for the next iteration
        driver.find_element(By.ID, "user-name").clear()
        driver.find_element(By.ID, "password").clear()
